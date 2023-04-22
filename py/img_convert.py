from google.colab import drive
drive.mount('/content/drive')

!pip install openpyxl
!pip install pillow

import os
from openpyxl import load_workbook
from PIL import ImageGrab

# Excelファイルのディレクトリとファイル名を指定
load = '/content/drive/MyDrive/Documents/sitemap.xlsx'

# 画像の保存先とファイル名を指定
save = '/content/drive/MyDrive/Image/sitemap.png'

# Excelファイルをロード
wb = load_workbook(filename=load)

# シートを取得
sheet = wb.active

# 最終行を取得
last_row = sheet.max_row

# セル範囲を指定
cell_range = f'A1:N{last_row}'

# セル範囲を画像に変換
img = sheet.sheet_to_png(range_string=cell_range)

# 画像を保存
img.save(save, 'PNG')
